import os

import numpy as np
import openpyxl
from tqdm import tqdm

from src.utils.cell_features import get_cell_features_xlsx, feature_order, parse_table_range


class SpreadsheetReader:
    def __init__(self, width, height):
        self._num_cell_features = 17
        self._map_width = width
        self._map_height = height
        self._empty_cell = np.array([1.0, 0.0, 0.0, 0.0, 0.0, 1.0, 1.0, 1.0, 1.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0])

    def read_spreadsheet(self, file_path, sheet_name, tables_area):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        min_row = 1
        max_row = min(ws.max_row, self._map_height)
        min_col = 1
        max_col = min(ws.max_column, self._map_width)

        sheet_tensor = np.zeros((max_row, max_col, self._num_cell_features), dtype=np.float32)

        for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)):
            for col_idx, cell in enumerate(row):
                features = get_cell_features_xlsx(cell)
                feature_matrix = np.array([float(features[key]) for key in feature_order], dtype=np.float32)
                sheet_tensor[row_idx, col_idx] = feature_matrix


        tables_features, background_features = self.separate_tables_data(sheet_tensor, tables_area)

        return tables_features, background_features

    def separate_tables_data(self, sheet_tensor, tables_area):
        background_features = sheet_tensor.copy()
        H, W, _ = background_features.shape
        tables_features = []
        for table_area in tables_area:
            min_col, min_row, max_col, max_row = parse_table_range(table_area)
            if min_col >= W or min_row >= H:
                continue
            table_slice = sheet_tensor[min_row:max_row, min_col:max_col].copy()
            h, w, c = table_slice.shape
            if h == 0 or w == 0:
                raise ValueError(f"Table slice has zero height or width: shape {table_slice.shape}")
            tables_features.append(table_slice)
            height, width = table_slice.shape[:2]

            sheet_tensor[min_row:max_row, min_col:max_col] = np.tile(self._empty_cell, (height, width, 1))

        return tables_features, background_features

    def load_dataset_maps(self, labels_df, data_folder):
        """
        Extracts feature maps from a labeled DataFrame using a custom feature extraction function.

        Args:
            labels_df (pd.DataFrame): DataFrame with columns ['sheet_name', 'file_path', 'table_region']
            data_folder (str): Root folder where spreadsheet files are located

        Returns:
            List[Tensor]: A list of feature maps extracted from the dataset
        """
        tables,  backgrounds = [], []

        for _, row in tqdm(labels_df.iterrows(), total=len(labels_df), desc="Processing sheets"):
            sheet_name = row['sheet_name']
            file_path = os.path.join(data_folder, row['file_path'])
            print(file_path)
            tables_features, background_features = self.read_spreadsheet(file_path, sheet_name, row['table_region'])
            tables.extend(tables_features)
            background_features = self.pad_feature_map(background_features)
            backgrounds.append(background_features)

        return tables,  backgrounds

    def pad_feature_map(self, feature_map):
        h, w, d = feature_map.shape

        # Clip dimensions if they exceed max limits
        clipped_h = min(h, self._map_height)
        clipped_w = min(w, self._map_width)

        # Optionally log or warn if clipping occurs
        if h > self._map_height or w > self._map_width:
            print(f"Warning: Feature map clipped from ({h}, {w}) to ({clipped_h}, {clipped_w})")

        # Create padded map filled with _empty_cell values
        resized_map = np.tile(self._empty_cell, (self._map_height, self._map_width, 1))

        # Copy the valid portion of the original map into the top-left corner
        resized_map[:clipped_h, :clipped_w, :] = feature_map[:clipped_h, :clipped_w, :]

        return resized_map

