import os

import pandas as pd
import openpyxl
import torch
from openpyxl.utils import range_boundaries




def get_cell_features_xlsx(cur_cell):
    cell_features = {
        "coordinate": cur_cell.coordinate,
        "is_empty": cur_cell.value is None,
        "is_string": cur_cell.data_type in ["s", "str"],
        "is_merged": type(cur_cell).__name__ == "MergedCell",
        "is_bold": cur_cell.font.b or False,
        "is_italic": cur_cell.font.i or False,
        "left_border": cur_cell.border.left is not None,
        "right_border": cur_cell.border.right is not None,
        "top_border": cur_cell.border.top is not None,
        "bottom_border": cur_cell.border.bottom is not None,
        "is_filled": cur_cell.fill.patternType is not None,
        "horizontal_alignment": cur_cell.alignment.horizontal is not None,
        "left_horizontal_alignment": cur_cell.alignment.horizontal == "left",
        "right_horizontal_alignment": cur_cell.alignment.horizontal == "right",
        "center_horizontal_alignment": cur_cell.alignment.horizontal == "center",
        "wrapped_text": cur_cell.alignment.wrapText or False,
        "indent": cur_cell.alignment.indent != 0,
        "formula": cur_cell.data_type == "f",
    }
    return cell_features


def get_table_features(file_path, sheet_name) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    # Determine the actual data range
    min_row = 1
    max_row = ws.max_row
    min_col = 1
    max_col = ws.max_column



    data = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            data.append(get_cell_features_xlsx(cell))

    result_df = pd.DataFrame(data)
    return result_df

feature_order = [
    'is_empty',
    'is_string',
    'is_merged',
    'is_bold',
    'is_italic',
    'left_border',
    'right_border',
    'top_border',
    'bottom_border',
    'is_filled',
    'horizontal_alignment',
    'left_horizontal_alignment',
    'right_horizontal_alignment',
    'center_horizontal_alignment',
    'wrapped_text',
    'indent',
    'formula'
]


def get_table_features2(file_path, sheet_name, table_area) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    # Determine the actual data range
    # min_row = 1
    # max_row = ws.max_row
    # min_col = 1
    # max_col = ws.max_column

    num_cell_features = 17
    #min_col, min_row, max_col, max_row = range_boundaries(table_area)
    min_col, min_row, max_col, max_row = parse_table_range(table_area)
    sheet_tensor = torch.zeros((max_row - min_row + 1, max_col - min_col + 1, num_cell_features), dtype=torch.float32)

    for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)):
        for col_idx, cell in enumerate(row):
            features = get_cell_features_xlsx(cell)
            feature_matrix = torch.tensor([float(features[key]) for key in feature_order],dtype=torch.float32)
            sheet_tensor[row_idx, col_idx] = feature_matrix

    return sheet_tensor

# def get_spreadsheet_features(file_path, sheet_name) -> pd.DataFrame:
#     wb = openpyxl.load_workbook(file_path)
#     ws = wb[sheet_name]
#
#     # Determine the actual data range
#     min_row = 1
#     max_row = ws.max_row
#     min_col = 1
#     max_col = ws.max_column
#
#     num_cell_features = 17
#
#     sheet_tensor = torch.zeros((max_row - min_row + 1, max_col - min_col + 1, num_cell_features), dtype=torch.float32)
#
#     for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)):
#         for col_idx, cell in enumerate(row):
#             features = get_cell_features_xlsx(cell)
#             feature_matrix = torch.tensor([float(features[key]) for key in feature_order], dtype=torch.float32)
#             sheet_tensor[row_idx, col_idx] = feature_matrix
#     return sheet_tensor

# def get_tables_features(sheet_tensor, table_areas) -> list:
#     tables = []
#     for table_area in table_areas:
#         min_col, min_row, max_col, max_row = parse_table_range(table_area)
#         table = sheet_tensor[min_col:max_col,min_row:max_row].copy()
#         tables.append(table)
#
#     return tables

# def remove_tables_from_spreadsheet(sheet_tensor, table_areas, device):
#     base_vector = torch.tensor([
#         1.0, 0.0, 0.0, 0.0, 0.0, 1.0, 1.0, 1.0, 1.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
#     ], device=device)
#
#     for table_area in table_areas:
#         min_col, min_row, max_col, max_row = parse_table_range(table_area)
#         sheet_tensor[ min_row:max_row + 1, min_col:max_col + 1, :] = base_vector

def generate_feature_tensor(H, W, device):
    """
    Generate a (H, W, 17) tensor where each cell contains the same predefined 17D vector.

    Args:
        H (int): Height of the output tensor.
        W (int): Width of the output tensor.

    Returns:
        torch.Tensor: A tensor of shape (H, W, 17)
    """
    # Define the 17D feature vector
    base_vector = torch.tensor([
        1.0, 0.0, 0.0, 0.0, 0.0, 1.0, 1.0, 1.0, 1.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
    ], device=device)

    # Repeat it H * W times and reshape to (H, W, 17)
    return base_vector.repeat(H * W, 1).view(H, W, 17)


def extract_feature_maps_from_labels(labels_df, data_folder):
    """
    Extracts feature maps from a labeled DataFrame using a custom feature extraction function.

    Args:
        labels_df (pd.DataFrame): DataFrame with columns ['sheet_name', 'file_path', 'table_region']
        data_folder (str): Root folder where spreadsheet files are located

    Returns:
        List[Tensor]: A list of feature maps extracted from the dataset
    """
    feature_maps = []

    for _, row in labels_df.iterrows():
        sheet_name = row['sheet_name']
        file_path = os.path.join(data_folder, row['file_path'])

        for table_area in row['table_region']:
            features_map = get_table_features2(file_path, sheet_name, table_area)
            feature_maps.append(features_map)

    return feature_maps

def parse_table_range(table_range):
    # Convert table range string to numerical coordinates
    # Placeholder for actual implementation
    min_col, min_row, max_col, max_row = range_boundaries(table_range)
    x_min = min_col - 1
    y_min = min_row - 1
    x_max = max_col - 1
    y_max = max_row - 1

    if x_min >= x_max or y_min >= y_max:
        raise ValueError(
            f"Table range '{table_range}' has zero or negative size: "
            f"(columns {x_min}–{x_max}, rows {y_min}–{y_max})"
        )
    return [x_min, y_min, x_max, y_max]
